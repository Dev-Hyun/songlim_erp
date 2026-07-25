from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.audit import log_action
from app.database import get_db
from app.models import (
    Equipment,
    GradeMaster,
    Hospital,
    HospitalProfile,
    SalesNote,
    SupplyCatalog,
    SupplyCategoryAccess,
    SupplyFavorite,
    SupplyOrder,
    SupplyOrderItem,
    SupplyPriceOverride,
    User,
)
from app.routers.auth import require_staff, require_user

router = APIRouter(prefix="/api/supply", tags=["supply"])


async def require_hospital(user: User = Depends(require_user)) -> User:
    if user.role != "hospital" or not user.hospital_profile_id:
        raise HTTPException(status_code=403, detail="병원 계정만 이용할 수 있습니다")
    return user


async def _get_hospital_profile(db: AsyncSession, hospital_profile_id: int) -> HospitalProfile:
    hp = (await db.execute(select(HospitalProfile).where(HospitalProfile.id == hospital_profile_id))).scalar_one_or_none()
    if not hp:
        raise HTTPException(status_code=404, detail="병원 정보를 찾을 수 없습니다")
    return hp


@router.get("/catalog")
async def list_catalog(
    db: AsyncSession = Depends(get_db),
    category: Optional[str] = None,
    user: User = Depends(require_hospital),
):
    hp = await _get_hospital_profile(db, user.hospital_profile_id)
    q = select(SupplyCatalog).where(SupplyCatalog.is_active.is_(True)).order_by(SupplyCatalog.category, SupplyCatalog.sub_category, SupplyCatalog.sort_order)
    if category:
        q = q.where(SupplyCatalog.category == category)
    items = (await db.execute(q)).scalars().all()

    access_rows = (await db.execute(select(SupplyCategoryAccess))).scalars().all()
    restricted = {r.category for r in access_rows}
    allowed = {r.category for r in access_rows if r.hospital_type == hp.hospital_type}

    overrides = {
        o.catalog_id: o.override_price
        for o in (
            await db.execute(select(SupplyPriceOverride).where(SupplyPriceOverride.hospital_profile_id == hp.id))
        ).scalars().all()
    }
    favorite_ids = {
        f.catalog_id
        for f in (
            await db.execute(select(SupplyFavorite).where(SupplyFavorite.hospital_profile_id == hp.id))
        ).scalars().all()
    }

    result = []
    for it in items:
        if it.category in restricted and it.category not in allowed:
            continue
        result.append({
            "id": it.id, "name": it.name, "spec": it.spec, "manufacturer": it.manufacturer,
            "category": it.category, "sub_category": it.sub_category,
            "unit": it.unit,
            "price": overrides.get(it.id, it.unit_price), "base_price": it.unit_price,
            "has_special_price": it.id in overrides,
            "description": it.description, "image_key": it.image_key,
            "is_favorite": it.id in favorite_ids,
        })
    return result


@router.get("/categories")
async def list_categories(db: AsyncSession = Depends(get_db), user: User = Depends(require_hospital)):
    hp = await _get_hospital_profile(db, user.hospital_profile_id)
    items = (await db.execute(select(SupplyCatalog.category).where(SupplyCatalog.is_active.is_(True)))).scalars().all()
    access_rows = (await db.execute(select(SupplyCategoryAccess))).scalars().all()
    restricted = {r.category for r in access_rows}
    allowed = {r.category for r in access_rows if r.hospital_type == hp.hospital_type}

    counts: dict[str, int] = {}
    for cat in items:
        if cat in restricted and cat not in allowed:
            continue
        counts[cat] = counts.get(cat, 0) + 1
    return [{"category": c, "count": n} for c, n in sorted(counts.items())]


@router.get("/subcategories")
async def list_subcategories(category: str, db: AsyncSession = Depends(get_db), user: User = Depends(require_hospital)):
    """특정 대분류(category) 안의 소분류별 품목 수."""
    rows = (
        await db.execute(
            select(SupplyCatalog.sub_category)
            .where(SupplyCatalog.is_active.is_(True), SupplyCatalog.category == category, SupplyCatalog.sub_category.is_not(None))
        )
    ).scalars().all()
    counts: dict[str, int] = {}
    for sc in rows:
        counts[sc] = counts.get(sc, 0) + 1
    return [{"sub_category": s, "count": n} for s, n in sorted(counts.items())]


class FavoriteIn(BaseModel):
    catalog_id: int


@router.post("/favorites")
async def add_favorite(payload: FavoriteIn, db: AsyncSession = Depends(get_db), user: User = Depends(require_hospital)):
    existing = (
        await db.execute(
            select(SupplyFavorite).where(
                SupplyFavorite.catalog_id == payload.catalog_id, SupplyFavorite.hospital_profile_id == user.hospital_profile_id
            )
        )
    ).scalar_one_or_none()
    if existing:
        return {"id": existing.id}
    f = SupplyFavorite(catalog_id=payload.catalog_id, hospital_profile_id=user.hospital_profile_id)
    db.add(f)
    await db.commit()
    await db.refresh(f)
    return {"id": f.id}


@router.delete("/favorites/{catalog_id}")
async def remove_favorite(catalog_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(require_hospital)):
    existing = (
        await db.execute(
            select(SupplyFavorite).where(
                SupplyFavorite.catalog_id == catalog_id, SupplyFavorite.hospital_profile_id == user.hospital_profile_id
            )
        )
    ).scalar_one_or_none()
    if existing:
        await db.delete(existing)
        await db.commit()
    return {"ok": True}


class OrderItemIn(BaseModel):
    catalog_id: int
    qty: int = 1


class OrderCreateIn(BaseModel):
    items: list[OrderItemIn]
    order_request: Optional[str] = None


@router.post("/orders")
async def create_order(payload: OrderCreateIn, db: AsyncSession = Depends(get_db), user: User = Depends(require_hospital)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="담긴 품목이 없습니다")
    hp = await _get_hospital_profile(db, user.hospital_profile_id)

    overrides = {
        o.catalog_id: o.override_price
        for o in (
            await db.execute(select(SupplyPriceOverride).where(SupplyPriceOverride.hospital_profile_id == hp.id))
        ).scalars().all()
    }

    order_items: list[SupplyOrderItem] = []
    subtotal = 0
    for it in payload.items:
        catalog = (await db.execute(select(SupplyCatalog).where(SupplyCatalog.id == it.catalog_id))).scalar_one_or_none()
        if not catalog:
            continue
        price = overrides.get(catalog.id, catalog.unit_price)
        line_total = price * it.qty
        subtotal += line_total
        order_items.append(SupplyOrderItem(
            catalog_id=catalog.id, name_snapshot=catalog.name,
            manufacturer_snapshot=catalog.manufacturer, spec_snapshot=catalog.spec,
            unit_snapshot=catalog.unit,
            unit_price_snapshot=price, qty=it.qty, subtotal=line_total,
        ))

    discount_rate = None
    gift_note = None
    if hp.discount_grade_code:
        g = (await db.execute(select(GradeMaster).where(GradeMaster.grade_code == hp.discount_grade_code))).scalar_one_or_none()
        discount_rate = g.discount_rate if g else None
    if hp.gift_grade_code:
        g = (await db.execute(select(GradeMaster).where(GradeMaster.grade_code == hp.gift_grade_code))).scalar_one_or_none()
        gift_note = g.gift_policy_note if g else None

    total = subtotal
    if discount_rate:
        total = round(subtotal * (1 - discount_rate / 100))

    order = SupplyOrder(
        hospital_profile_id=hp.id, ordered_by=user.id, total_amount=total,
        discount_rate_applied=discount_rate, gift_note=gift_note,
        order_request=(payload.order_request or None),
    )
    order.items = order_items
    db.add(order)
    await db.flush()

    await db.commit()
    await db.refresh(order)
    return {"id": order.id, "total_amount": order.total_amount}


def _serialize_order(o: SupplyOrder) -> dict:
    return {
        "id": o.id, "status": o.status, "tracking_number": o.tracking_number,
        "total_amount": o.total_amount, "discount_rate_applied": o.discount_rate_applied,
        "gift_note": o.gift_note, "tax_invoice_status": o.tax_invoice_status,
        "order_request": o.order_request,
        "created_at": o.created_at, "updated_at": o.updated_at,
        "items": [
            {"id": i.id, "catalog_id": i.catalog_id, "name": i.name_snapshot,
             "manufacturer": i.manufacturer_snapshot, "spec": i.spec_snapshot,
             "unit": i.unit_snapshot,
             "unit_price": i.unit_price_snapshot, "qty": i.qty, "subtotal": i.subtotal}
            for i in o.items
        ],
    }


@router.get("/orders")
async def list_my_orders(
    db: AsyncSession = Depends(get_db),
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: User = Depends(require_hospital),
):
    q = select(SupplyOrder).options(selectinload(SupplyOrder.items)).where(SupplyOrder.hospital_profile_id == user.hospital_profile_id)
    if status:
        q = q.where(SupplyOrder.status == status)
    if date_from:
        q = q.where(SupplyOrder.created_at >= date_from)
    if date_to:
        q = q.where(SupplyOrder.created_at <= date_to)
    rows = (await db.execute(q.order_by(SupplyOrder.created_at.desc()))).scalars().all()
    return [_serialize_order(o) for o in rows]


@router.get("/orders/{oid}")
async def get_my_order(oid: int, db: AsyncSession = Depends(get_db), user: User = Depends(require_hospital)):
    o = (await db.execute(select(SupplyOrder).options(selectinload(SupplyOrder.items)).where(SupplyOrder.id == oid))).scalar_one_or_none()
    if not o or o.hospital_profile_id != user.hospital_profile_id:
        raise HTTPException(status_code=404, detail="발주 내역을 찾을 수 없습니다")
    return _serialize_order(o)


@router.get("/orders/{oid}/reorder-items")
async def reorder_items(oid: int, db: AsyncSession = Depends(get_db), user: User = Depends(require_hospital)):
    """기존 발주의 품목/수량을 그대로 프론트 장바구니에 채워넣기 위한 프리필 데이터."""
    o = (await db.execute(select(SupplyOrder).options(selectinload(SupplyOrder.items)).where(SupplyOrder.id == oid))).scalar_one_or_none()
    if not o or o.hospital_profile_id != user.hospital_profile_id:
        raise HTTPException(status_code=404, detail="발주 내역을 찾을 수 없습니다")
    return [{"catalog_id": i.catalog_id, "name": i.name_snapshot, "qty": i.qty} for i in o.items if i.catalog_id]


# ────────────────────────────────────────────────────────
# 관리자(송림 직원)측 — 카탈로그/카테고리권한/병원별 등급·가격오버라이드/발주관리
# ────────────────────────────────────────────────────────
class CatalogIn(BaseModel):
    code: Optional[str] = None
    name: str
    manufacturer: Optional[str] = None
    spec: Optional[str] = None
    category: str = "기타"
    sub_category: Optional[str] = None
    unit: str = "개"  # 자유 입력, 예: "100입/1box"
    unit_price: int = 0
    description: Optional[str] = None
    image_key: Optional[str] = None
    sort_order: int = 0
    is_active: bool = True


def _serialize_catalog(c: SupplyCatalog) -> dict:
    return {
        "id": c.id, "code": c.code, "name": c.name, "manufacturer": c.manufacturer,
        "spec": c.spec, "category": c.category, "sub_category": c.sub_category,
        "unit": c.unit,
        "unit_price": c.unit_price, "description": c.description, "image_key": c.image_key,
        "sort_order": c.sort_order, "is_active": c.is_active,
    }


@router.get("/admin/catalog")
async def admin_list_catalog(db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    rows = (await db.execute(select(SupplyCatalog).order_by(SupplyCatalog.category, SupplyCatalog.sort_order))).scalars().all()
    return [_serialize_catalog(c) for c in rows]


@router.post("/admin/catalog")
async def admin_create_catalog(payload: CatalogIn, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    c = SupplyCatalog(**payload.model_dump())
    db.add(c)
    await db.commit()
    await db.refresh(c)
    return _serialize_catalog(c)


@router.patch("/admin/catalog/{cid}")
async def admin_update_catalog(cid: int, payload: CatalogIn, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    c = (await db.execute(select(SupplyCatalog).where(SupplyCatalog.id == cid))).scalar_one_or_none()
    if not c:
        raise HTTPException(status_code=404, detail="품목을 찾을 수 없습니다")
    for k, v in payload.model_dump().items():
        setattr(c, k, v)
    await db.commit()
    return _serialize_catalog(c)


@router.delete("/admin/catalog/{cid}")
async def admin_delete_catalog(cid: int, db: AsyncSession = Depends(get_db), actor: User = Depends(require_staff)):
    c = (await db.execute(select(SupplyCatalog).where(SupplyCatalog.id == cid))).scalar_one_or_none()
    if c:
        log_action(db, actor, "post_delete", "supply_catalog", cid, detail=c.name)
        await db.delete(c)
        await db.commit()
    return {"ok": True}


CATALOG_XLSX_HEADERS = ["코드", "품목명", "제조사", "카테고리", "소분류", "단위", "기본금액", "노출여부"]


@router.get("/admin/catalog/export")
async def admin_export_catalog(db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    import io
    from urllib.parse import quote

    import openpyxl
    from fastapi.responses import StreamingResponse

    rows = (await db.execute(select(SupplyCatalog).order_by(SupplyCatalog.category, SupplyCatalog.sort_order))).scalars().all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "소모품 카탈로그"
    ws.append(CATALOG_XLSX_HEADERS)
    for c in rows:
        ws.append([c.code or "", c.name, c.manufacturer or "", c.category, c.sub_category or "", c.unit, c.unit_price, "O" if c.is_active else "X"])
    for i, w in enumerate([12, 28, 16, 14, 14, 16, 12, 10], start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "소모품_카탈로그.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


@router.post("/admin/catalog/import")
async def admin_import_catalog(file: UploadFile = File(...), db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    """엑셀 업로드로 카탈로그 일괄 추가. 이미 등록된 품목명(중복)은 건너뛰고 신규 항목만 추가한다."""
    import io

    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    existing_names = {n for (n,) in (await db.execute(select(SupplyCatalog.name))).all()}
    added = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2 or not row[1]:
            continue
        code = str(row[0]).strip() if row[0] else None
        name = str(row[1]).strip()
        if not name or name in existing_names:
            skipped += 1
            continue
        manufacturer = str(row[2]).strip() if len(row) > 2 and row[2] else None
        category = str(row[3]).strip() if len(row) > 3 and row[3] else "기타"
        sub_category = str(row[4]).strip() if len(row) > 4 and row[4] else None
        unit = str(row[5]).strip() if len(row) > 5 and row[5] else "개"
        unit_price = int(row[6]) if len(row) > 6 and row[6] else 0
        is_active = not (len(row) > 7 and str(row[7]).strip().upper() == "X")
        db.add(SupplyCatalog(
            code=code, name=name, manufacturer=manufacturer, category=category, sub_category=sub_category,
            unit=unit, unit_price=unit_price, is_active=is_active,
        ))
        existing_names.add(name)
        added += 1
    await db.commit()
    return {"added": added, "skipped": skipped}


class CatalogReorderIn(BaseModel):
    ordered_ids: list[int]


@router.post("/admin/catalog/reorder")
async def admin_reorder_catalog(payload: CatalogReorderIn, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    """드래그앤드롭으로 정렬한 순서를 그대로 sort_order에 반영."""
    rows = (await db.execute(select(SupplyCatalog).where(SupplyCatalog.id.in_(payload.ordered_ids)))).scalars().all()
    by_id = {c.id: c for c in rows}
    for i, cid in enumerate(payload.ordered_ids):
        if cid in by_id:
            by_id[cid].sort_order = i
    await db.commit()
    return {"ok": True}


class CategoryAccessIn(BaseModel):
    category: str
    hospital_type: str


@router.get("/admin/category-access")
async def admin_list_category_access(db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    rows = (await db.execute(select(SupplyCategoryAccess))).scalars().all()
    return [{"id": r.id, "category": r.category, "hospital_type": r.hospital_type} for r in rows]


@router.post("/admin/category-access")
async def admin_add_category_access(payload: CategoryAccessIn, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    r = SupplyCategoryAccess(**payload.model_dump())
    db.add(r)
    await db.commit()
    await db.refresh(r)
    return {"id": r.id}


@router.delete("/admin/category-access/{rid}")
async def admin_remove_category_access(rid: int, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    r = (await db.execute(select(SupplyCategoryAccess).where(SupplyCategoryAccess.id == rid))).scalar_one_or_none()
    if r:
        await db.delete(r)
        await db.commit()
    return {"ok": True}


@router.get("/admin/hospitals")
async def admin_list_hospitals(db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    rows = (await db.execute(select(HospitalProfile))).scalars().all()
    return [
        {"id": h.id, "hospital_name": h.hospital_name, "hospital_type": h.hospital_type,
         "discount_grade_code": h.discount_grade_code, "gift_grade_code": h.gift_grade_code}
        for h in rows
    ]


@router.get("/admin/hospitals/{hid}/detail")
async def admin_get_hospital_detail(hid: int, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    """회원가입시 입력한 병원 정보 전체 + (영업지도와 연동된 경우) 장비 보유 현황 + 영업노트."""
    hp = await _get_hospital_profile(db, hid)

    contact_rows = (await db.execute(select(User).where(User.hospital_profile_id == hid))).scalars().all()
    contacts = [
        {"username": u.username, "display_name": u.display_name, "phone": u.phone, "email": u.email}
        for u in contact_rows
    ]

    matched = (await db.execute(select(Hospital).where(Hospital.hospital_profile_id == hid))).scalar_one_or_none()
    equipment = []
    notes = []
    if matched:
        eq_rows = (await db.execute(select(Equipment).where(Equipment.hospital_id == matched.id))).scalars().all()
        equipment = [
            {"id": e.id, "category": e.category, "year": e.year, "manufacturer": e.manufacturer,
             "model": e.model, "eq_count": e.eq_count, "source": e.source}
            for e in eq_rows
        ]
        note_rows = (
            await db.execute(
                select(SalesNote, User.display_name, User.username)
                .join(User, User.id == SalesNote.user_id)
                .where(SalesNote.hospital_id == matched.id)
                .order_by(SalesNote.visit_date.desc())
            )
        ).all()
        notes = [
            {"id": n.id, "visit_date": n.visit_date.isoformat() if n.visit_date else None,
             "content": n.content, "created_by_name": dn or un}
            for n, dn, un in note_rows
        ]

    return {
        "id": hp.id,
        "hospital_name": hp.hospital_name,
        "hospital_type": hp.hospital_type,
        "hospital_dept": hp.hospital_dept,
        "hospital_address": hp.hospital_address,
        "hospital_tel": hp.hospital_tel,
        "business_reg_no": hp.business_reg_no,
        "ceo_name": hp.ceo_name,
        "ceo_phone": hp.ceo_phone,
        "contacts": contacts,
        "discount_grade_code": hp.discount_grade_code,
        "gift_grade_code": hp.gift_grade_code,
        "matched_hospital_id": matched.id if matched else None,
        "equipment": equipment,
        "sales_notes": notes,
    }


class HospitalGradeIn(BaseModel):
    discount_grade_code: Optional[str] = None
    gift_grade_code: Optional[str] = None


@router.patch("/admin/hospitals/{hid}/grades")
async def admin_set_hospital_grades(hid: int, payload: HospitalGradeIn, db: AsyncSession = Depends(get_db), actor: User = Depends(require_staff)):
    hp = await _get_hospital_profile(db, hid)
    hp.discount_grade_code = payload.discount_grade_code
    hp.gift_grade_code = payload.gift_grade_code
    log_action(db, actor, "price_change", "hospital_grade", hid,
               detail=f"discount={payload.discount_grade_code}, gift={payload.gift_grade_code}")
    await db.commit()
    return {"ok": True}


class PriceOverrideIn(BaseModel):
    catalog_id: int
    hospital_profile_id: int
    override_price: int


@router.get("/admin/price-overrides")
async def admin_list_price_overrides(db: AsyncSession = Depends(get_db), hospital_profile_id: Optional[int] = None, _: User = Depends(require_staff)):
    q = select(SupplyPriceOverride)
    if hospital_profile_id:
        q = q.where(SupplyPriceOverride.hospital_profile_id == hospital_profile_id)
    rows = (await db.execute(q)).scalars().all()
    return [{"id": o.id, "catalog_id": o.catalog_id, "hospital_profile_id": o.hospital_profile_id, "override_price": o.override_price} for o in rows]


@router.post("/admin/price-overrides")
async def admin_upsert_price_override(payload: PriceOverrideIn, db: AsyncSession = Depends(get_db), actor: User = Depends(require_staff)):
    existing = (
        await db.execute(
            select(SupplyPriceOverride).where(
                SupplyPriceOverride.catalog_id == payload.catalog_id,
                SupplyPriceOverride.hospital_profile_id == payload.hospital_profile_id,
            )
        )
    ).scalar_one_or_none()
    log_action(db, actor, "price_change", "price_override",
               f"catalog={payload.catalog_id},hospital={payload.hospital_profile_id}",
               detail=f"override_price={payload.override_price}")
    if existing:
        existing.override_price = payload.override_price
        await db.commit()
        return {"id": existing.id}
    o = SupplyPriceOverride(**payload.model_dump())
    db.add(o)
    await db.commit()
    await db.refresh(o)
    return {"id": o.id}


@router.delete("/admin/price-overrides/{oid}")
async def admin_delete_price_override(oid: int, db: AsyncSession = Depends(get_db), actor: User = Depends(require_staff)):
    o = (await db.execute(select(SupplyPriceOverride).where(SupplyPriceOverride.id == oid))).scalar_one_or_none()
    if o:
        log_action(db, actor, "price_change", "price_override", oid, detail="삭제")
        await db.delete(o)
        await db.commit()
    return {"ok": True}


@router.get("/admin/orders")
async def admin_list_orders(
    db: AsyncSession = Depends(get_db),
    status: Optional[str] = None,
    hospital_profile_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    _: User = Depends(require_staff),
):
    q = select(SupplyOrder).options(selectinload(SupplyOrder.items))
    if status:
        q = q.where(SupplyOrder.status == status)
    if hospital_profile_id:
        q = q.where(SupplyOrder.hospital_profile_id == hospital_profile_id)
    if date_from:
        q = q.where(SupplyOrder.created_at >= date_from)
    if date_to:
        q = q.where(SupplyOrder.created_at <= date_to)
    rows = (await db.execute(q.order_by(SupplyOrder.created_at.desc()))).scalars().all()
    result = []
    for o in rows:
        hp = (await db.execute(select(HospitalProfile).where(HospitalProfile.id == o.hospital_profile_id))).scalar_one_or_none()
        d = _serialize_order(o)
        d["hospital_name"] = hp.hospital_name if hp else None
        d["hospital_profile_id"] = o.hospital_profile_id
        result.append(d)
    return result


@router.get("/admin/orders/{oid}")
async def admin_get_order(oid: int, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    o = (await db.execute(select(SupplyOrder).options(selectinload(SupplyOrder.items)).where(SupplyOrder.id == oid))).scalar_one_or_none()
    if not o:
        raise HTTPException(status_code=404, detail="발주 내역을 찾을 수 없습니다")
    hp = (await db.execute(select(HospitalProfile).where(HospitalProfile.id == o.hospital_profile_id))).scalar_one_or_none()
    d = _serialize_order(o)
    d["hospital_name"] = hp.hospital_name if hp else None
    return d


class OrderStatusIn(BaseModel):
    status: str


@router.patch("/admin/orders/{oid}/status")
async def admin_update_order_status(oid: int, payload: OrderStatusIn, db: AsyncSession = Depends(get_db), actor: User = Depends(require_staff)):
    if payload.status not in ("접수", "출고", "배송완료", "직납출고", "직납완료"):
        raise HTTPException(status_code=400, detail="잘못된 상태값입니다")
    o = (await db.execute(select(SupplyOrder).where(SupplyOrder.id == oid))).scalar_one_or_none()
    if not o:
        raise HTTPException(status_code=404, detail="발주 내역을 찾을 수 없습니다")
    log_action(db, actor, "order_status_change", "supply_order", oid, detail=f"{o.status} → {payload.status}")
    o.status = payload.status
    await db.commit()
    return {"ok": True}


class TrackingIn(BaseModel):
    tracking_number: str


@router.patch("/admin/orders/{oid}/tracking")
async def admin_set_tracking(oid: int, payload: TrackingIn, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    o = (await db.execute(select(SupplyOrder).where(SupplyOrder.id == oid))).scalar_one_or_none()
    if not o:
        raise HTTPException(status_code=404, detail="발주 내역을 찾을 수 없습니다")
    o.tracking_number = payload.tracking_number
    await db.commit()
    return {"ok": True}


class TaxInvoiceIn(BaseModel):
    tax_invoice_status: str


@router.patch("/admin/orders/{oid}/tax-invoice")
async def admin_set_tax_invoice_status(oid: int, payload: TaxInvoiceIn, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    """실제 국세청 전자세금계산서 발행 연동 전까지는 상태값만 수기로 기록 (ISSUE 참고)."""
    if payload.tax_invoice_status not in ("미발행", "발행요청", "발행완료"):
        raise HTTPException(status_code=400, detail="잘못된 상태값입니다")
    o = (await db.execute(select(SupplyOrder).where(SupplyOrder.id == oid))).scalar_one_or_none()
    if not o:
        raise HTTPException(status_code=404, detail="발주 내역을 찾을 수 없습니다")
    o.tax_invoice_status = payload.tax_invoice_status
    await db.commit()
    return {"ok": True}


