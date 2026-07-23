from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models import MileageLog, StaffProfile, User
from app.routers.auth import require_staff

router = APIRouter(prefix="/api", tags=["misc"])


# ────────────────────────────────────────────────────────
# 운행일지
# ────────────────────────────────────────────────────────
class MileageLogIn(BaseModel):
    log_date: str
    final_km: float
    prev_km: Optional[float] = None  # 미지정 시 직전 기록의 주행후 계기판거리로 자동 채움
    nonbiz_km: float = 0
    purpose: Optional[str] = ""
    note: Optional[str] = ""
    vehicle: Optional[str] = ""


@router.get("/mileage-logs")
async def list_mileage_logs(db: AsyncSession = Depends(get_db), user: User = Depends(require_staff)):
    rows = (
        await db.execute(select(MileageLog).where(MileageLog.user_id == user.id).order_by(MileageLog.log_date.desc()))
    ).scalars().all()
    # 차번호 자동저장: 이 사용자의 가장 최근 기록의 차량을 기본값으로 제공
    last_vehicle = rows[0].vehicle if rows else ""
    return {
        "last_vehicle": last_vehicle,
        "logs": [
            {"id": m.id, "log_date": m.log_date, "final_km": m.final_km, "prev_km": m.prev_km,
             "daily_km": m.daily_km, "nonbiz_km": m.nonbiz_km, "purpose": m.purpose, "note": m.note, "vehicle": m.vehicle}
            for m in rows
        ],
    }


async def _prev_km_for(db: AsyncSession, user_id: int, log_date: str) -> float:
    """해당 날짜 직전 기록의 주행후 계기판거리(final_km)를 전일 km로 사용 (레거시와 동일)."""
    row = (
        await db.execute(
            select(MileageLog.final_km)
            .where(MileageLog.user_id == user_id, MileageLog.log_date < log_date)
            .order_by(MileageLog.log_date.desc())
            .limit(1)
        )
    ).scalar_one_or_none()
    return row or 0.0


@router.post("/mileage-logs")
async def create_mileage_log(payload: MileageLogIn, db: AsyncSession = Depends(get_db), user: User = Depends(require_staff)):
    prev_km = payload.prev_km if payload.prev_km is not None else await _prev_km_for(db, user.id, payload.log_date)
    # 차번호 자동저장: 미입력 시 직전 기록의 차량 재사용
    vehicle = payload.vehicle
    if not vehicle:
        last = (
            await db.execute(
                select(MileageLog.vehicle).where(MileageLog.user_id == user.id).order_by(MileageLog.log_date.desc()).limit(1)
            )
        ).scalar_one_or_none()
        vehicle = last or ""
    m = MileageLog(
        user_id=user.id, log_date=payload.log_date, final_km=payload.final_km, prev_km=prev_km,
        daily_km=round(payload.final_km - prev_km, 1), nonbiz_km=payload.nonbiz_km,
        purpose=payload.purpose, note=payload.note, vehicle=vehicle,
        created_at=datetime.now(timezone.utc).isoformat(),
    )
    db.add(m)
    await db.commit()
    await db.refresh(m)
    return {"id": m.id}


@router.get("/mileage-logs/export")
async def export_mileage(year: int = 0, month: int = 0, db: AsyncSession = Depends(get_db), user: User = Depends(require_staff)):
    """국세청 업무용승용차 운행기록부 양식 Excel (레거시 /api/mileage/export 1:1 이식)."""
    import calendar as _cal
    import io
    from urllib.parse import quote

    import openpyxl
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    q = select(MileageLog).where(MileageLog.user_id == user.id)
    if year:
        q = q.where(MileageLog.log_date.like(f"{year}-%"))
    if month:
        q = q.where(MileageLog.log_date.like(f"%-{str(month).zfill(2)}-%"))
    rows = (await db.execute(q.order_by(MileageLog.log_date))).scalars().all()

    staff = (await db.execute(select(StaffProfile).where(StaffProfile.user_id == user.id))).scalar_one_or_none()
    dept = (staff.department if staff else "") or ""
    driver = user.display_name or user.username or ""
    vehicle = rows[0].vehicle if rows else ""

    if year and month:
        last_day = _cal.monthrange(year, month)[1]
        period = f"{year}-{month:02d}-01  ~  {year}-{month:02d}-{last_day:02d}"
    elif year:
        period = f"{year}-01-01  ~  {year}-12-31"
    else:
        period = ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "운행기록부"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="EEF2F7")
    sum_fill = PatternFill("solid", fgColor="FFF7E6")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    NCOL = 10

    def merge(r1, c1, r2, c2, value, *, bold=False, size=10, fill=None, align=center, font_color="000000"):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=value)
        cell.font = Font(name="맑은 고딕", bold=bold, size=size, color=font_color)
        cell.alignment = align
        if fill:
            cell.fill = fill
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(row=rr, column=cc).border = border

    merge(1, 1, 1, NCOL, "업무용승용차 운행기록부", bold=True, size=16)
    ws.row_dimensions[1].height = 30
    merge(2, 1, 2, 2, "과세기간", bold=True, fill=hdr_fill)
    merge(2, 3, 2, NCOL, period, align=left)
    merge(3, 1, 3, 2, "상호(법인명)", bold=True, fill=hdr_fill)
    merge(3, 3, 3, 5, "송림메디칼(주)", align=left)
    merge(3, 6, 3, 7, "사업자등록번호", bold=True, fill=hdr_fill)
    merge(3, 8, 3, NCOL, "", align=left)
    merge(4, 1, 4, 2, "차종", bold=True, fill=hdr_fill)
    merge(4, 3, 4, 5, "", align=left)
    merge(4, 6, 4, 7, "자동차등록번호", bold=True, fill=hdr_fill)
    merge(4, 8, 4, NCOL, vehicle, align=left)

    H = 5
    merge(H, 1, H + 1, 1, "①사용일자", bold=True, fill=hdr_fill)
    merge(H, 2, H + 1, 2, "②부서", bold=True, fill=hdr_fill)
    merge(H, 3, H + 1, 3, "③성명", bold=True, fill=hdr_fill)
    merge(H, 4, H + 1, 4, "④주행 전\n계기판거리(㎞)", bold=True, fill=hdr_fill)
    merge(H, 5, H + 1, 5, "⑤주행 후\n계기판거리(㎞)", bold=True, fill=hdr_fill)
    merge(H, 6, H + 1, 6, "⑥주행거리(㎞)\n(⑤-④)", bold=True, fill=hdr_fill)
    merge(H, 7, H, 8, "업무용 사용거리(㎞)", bold=True, fill=hdr_fill)
    merge(H + 1, 7, H + 1, 7, "⑦출퇴근용", bold=True, fill=hdr_fill)
    merge(H + 1, 8, H + 1, 8, "⑧일반업무용", bold=True, fill=hdr_fill)
    merge(H, 9, H + 1, 9, "⑨비업무용\n사용거리(㎞)", bold=True, fill=hdr_fill)
    merge(H, 10, H + 1, 10, "비고", bold=True, fill=hdr_fill)
    ws.row_dimensions[H].height = 20
    ws.row_dimensions[H + 1].height = 20

    r = H + 2
    tot_drive = tot_commute = tot_biz = tot_nonbiz = 0.0
    for rec in rows:
        daily = rec.daily_km or 0.0
        nonbiz = rec.nonbiz_km or 0.0
        biz = max(0.0, daily - nonbiz)
        vals = [(rec.log_date or "")[:10], dept, driver, rec.prev_km, rec.final_km, daily, 0.0, biz, nonbiz, rec.purpose or ""]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.border = border
            c.font = Font(name="맑은 고딕", size=10)
            if ci in (1, 2, 3):
                c.alignment = center
            elif ci == 10:
                c.alignment = left
            else:
                c.alignment = right
                c.number_format = "#,##0.0"
        tot_drive += daily
        tot_biz += biz
        tot_nonbiz += nonbiz
        r += 1

    merge(r, 1, r, 3, "합계", bold=True, fill=sum_fill)
    for ci, v in [(6, tot_drive), (7, tot_commute), (8, tot_biz), (9, tot_nonbiz)]:
        c = ws.cell(row=r, column=ci, value=v)
        c.border = border
        c.fill = sum_fill
        c.font = Font(name="맑은 고딕", bold=True, size=10)
        c.alignment = right
        c.number_format = "#,##0.0"
    for ci in (4, 5, 10):
        c = ws.cell(row=r, column=ci)
        c.border = border
        c.fill = sum_fill
    r += 2

    biz_total = tot_commute + tot_biz
    ratio = (biz_total / tot_drive * 100) if tot_drive else 0.0
    merge(r, 1, r, 4, "⑩ 과세기간 총주행거리(㎞)", bold=True, fill=hdr_fill, align=left)
    merge(r, 5, r, NCOL, f"{tot_drive:,.1f}", align=left, bold=True)
    r += 1
    merge(r, 1, r, 4, "⑪ 과세기간 업무용 사용거리(㎞) (⑦+⑧)", bold=True, fill=hdr_fill, align=left)
    merge(r, 5, r, NCOL, f"{biz_total:,.1f}", align=left, bold=True)
    r += 1
    merge(r, 1, r, 4, "⑫ 업무사용비율 (⑪/⑩)", bold=True, fill=hdr_fill, align=left)
    merge(r, 5, r, NCOL, f"{ratio:.1f} %", align=left, bold=True, font_color="C0392B")

    for i, w in enumerate([13, 9, 9, 12, 12, 12, 10, 11, 12, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    name_part = (f"_{year}" if year else "") + (f"_{str(month).zfill(2)}" if month else "")
    fname = f"업무용승용차_운행기록부{name_part}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


# ────────────────────────────────────────────────────────
# 입찰정보/의료소식은 app/routers/bids_news.py 에서 자동수집(G2B/D2B, RSS/스크래핑) 방식으로 이전됨
# ────────────────────────────────────────────────────────
