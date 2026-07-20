import math
import os
import xml.etree.ElementTree as ET
from datetime import date
from typing import Optional

import requests
from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from sqlalchemy import select, func, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models import Hospital, Equipment, SalesNote, User
from app.routers.auth import require_staff

router = APIRouter(prefix="/api", tags=["sales_map"])

HIRA_API_KEY = os.environ.get("HIRA_API_KEY", "")
HIRA_SEARCH_URL = "https://apis.data.go.kr/B551182/hospInfoServicev2/getHospBasisList"


def haversine(lat1, lng1, lat2, lng2):
    r = 6371
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp, dl = math.radians(lat2 - lat1), math.radians(lng2 - lng1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


# ────────────────────────────────────────────────────────
# 영업지도 조회
# 요구사항: 선택한 장비 대분류(us/xray/ct/mri)를 보유한 병원만 지도에 표시.
# 미보유 병원은 지도엔 안 나오되, 병원명 검색에는 나오고 항상 하단으로 정렬.
# 회원가입한 병원(hospital_profile_id 존재)은 회원 배지 표시.
# ────────────────────────────────────────────────────────
async def _query_hospitals(
    db: AsyncSession,
    lat: Optional[float],
    lng: Optional[float],
    radius_km: float,
    category: str,
    sido: Optional[str],
    sigungu: Optional[str],
    maker: Optional[str],
    model: Optional[str],
    name_search: Optional[str],
    map_only: bool,
    sort: str,
):
    q = select(Hospital)
    if sido:
        q = q.where(Hospital.sido == sido)
    if sigungu:
        q = q.where(Hospital.sigungu == sigungu)
    if name_search:
        q = q.where(Hospital.name.ilike(f"%{name_search}%"))
    # 지도 표시 모드 + 좌표 있음: 전국 59,253건 전체 스캔을 피하기 위해 RTREE 공간인덱스로 후보 id를 먼저 좁힌 뒤 조회
    # (레거시에서 이 인덱스가 빠져있던 게 영업지도 로딩이 느렸던 원인이었음 — 개발요청서 5-4/DB ERD 참고)
    if map_only and lat is not None and lng is not None and not name_search:
        lat_d = radius_km / 111.0
        lng_d = radius_km / (111.0 * math.cos(math.radians(lat)))
        rtree_ids = (
            await db.execute(
                text(
                    "SELECT id FROM hospitals_rtree WHERE min_lat <= :max_lat AND max_lat >= :min_lat "
                    "AND min_lng <= :max_lng AND max_lng >= :min_lng"
                ),
                {"min_lat": lat - lat_d, "max_lat": lat + lat_d, "min_lng": lng - lng_d, "max_lng": lng + lng_d},
            )
        ).scalars().all()
        if not rtree_ids:
            return []
        q = q.where(Hospital.id.in_(rtree_ids))
    hospitals = (await db.execute(q)).scalars().all()

    hosp_ids = [h.id for h in hospitals]
    eq_rows = []
    BATCH = 500  # SQLite 바인드변수 한도 대응 — 전국 검색처럼 hosp_ids가 많을 때 IN절을 분할
    for i in range(0, len(hosp_ids), BATCH):
        batch = hosp_ids[i:i + BATCH]
        eq_q = (
            select(Equipment)
            .where(Equipment.hospital_id.in_(batch), Equipment.category == category)
            .order_by(Equipment.hospital_id, Equipment.year.desc())
        )
        eq_rows.extend((await db.execute(eq_q)).scalars().all())

    eq_by_hosp: dict[int, list[Equipment]] = {}
    for e in eq_rows:
        eq_by_hosp.setdefault(e.hospital_id, []).append(e)
    if maker:
        eq_by_hosp = {
            hid: rows for hid, rows in eq_by_hosp.items()
            if any(r.manufacturer == maker for r in rows if r.year == rows[0].year)
        }
    if model:
        eq_by_hosp = {
            hid: rows for hid, rows in eq_by_hosp.items()
            if any(r.model == model for r in rows if r.year == rows[0].year)
        }

    # 장비 미보유 병원은 병원명 검색일 때만 노출한다. 시도/제조사/모델 필터(map_only=False)는
    # 반경 제한만 풀 뿐 "장비 보유 병원만 보여준다"는 규칙과는 무관하므로 map_only가 아니라
    # name_search 여부로 따로 판단해야 한다 (map_only에 묶으면 지역/제조사로 필터링할 때
    # 근처 장비 미보유 병원까지 섞여 나오는 버그가 생김).
    require_equipment = not name_search

    results = []
    for h in hospitals:
        eq = eq_by_hosp.get(h.id, [])
        has_equipment = bool(eq)

        if require_equipment and not has_equipment:
            continue

        dist_km = None
        if lat is not None and lng is not None and h.lat is not None and h.lng is not None:
            dist_km = round(haversine(lat, lng, h.lat, h.lng), 2)
            if map_only and dist_km > radius_km:
                continue

        latest = eq[0] if eq else None
        results.append({
            "id": h.id, "name": h.name, "type": h.type,
            "sido": h.sido, "sigungu": h.sigungu,
            "lat": h.lat, "lng": h.lng, "dist_km": dist_km,
            "is_member": h.hospital_profile_id is not None,
            "has_equipment": has_equipment,
            "current_maker": latest.manufacturer if latest else None,
            "current_model": latest.model if latest else None,
        })

    # 검색 결과는 항상 장비 보유 병원 상단, 미보유 하단
    if sort == "maker":
        results.sort(key=lambda x: (not x["has_equipment"], x["current_maker"] or "zzz", x["name"]))
    elif sort == "name":
        results.sort(key=lambda x: (not x["has_equipment"], x["name"]))
    else:
        results.sort(key=lambda x: (not x["has_equipment"], x["dist_km"] if x["dist_km"] is not None else 9e9, x["name"]))

    return results


@router.get("/hospitals")
async def get_hospitals(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_staff),
    lat: Optional[float] = None,
    lng: Optional[float] = None,
    radius_km: float = 3.0,
    category: str = "us",
    sido: Optional[str] = None,
    sigungu: Optional[str] = None,
    maker: Optional[str] = None,
    model: Optional[str] = None,
    name_search: Optional[str] = None,
    map_only: bool = True,  # True: 지도 표시용(장비 보유만), False: 검색용(전체 + 정렬)
    sort: str = "dist",
):
    results = await _query_hospitals(db, lat, lng, radius_km, category, sido, sigungu, maker, model, name_search, map_only, sort)
    return results[:500]


@router.get("/hospitals/export")
async def export_hospitals_excel(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_staff),
    lat: float = 37.5665,
    lng: float = 126.978,
    radius_km: float = 3.0,
    category: str = "us",
    sido: Optional[str] = None,
    sigungu: Optional[str] = None,
    maker: Optional[str] = None,
    model: Optional[str] = None,
    map_only: bool = True,
):
    """영업지도 목록 Excel 내보내기 (레거시 /api/hospitals/export 참고)."""
    import io
    from urllib.parse import quote

    import openpyxl
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    results = await _query_hospitals(db, lat, lng, radius_km, category, sido, sigungu, maker, model, None, map_only, "name")
    cat_label = {"us": "초음파", "xray": "엑스레이", "ct": "CT", "mri": "MRI", "bmd": "BMD", "carm": "C-Arm"}.get(category, category)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{cat_label} 영업지도"

    hdr_fill = PatternFill("solid", fgColor="1B64DA")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Border(*(Side(style="thin", color="D5DDE5"),) * 4)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cols = ["순위", "병원명", "시도", "시군구", "병원종별", f"현재 {cat_label}장비", "제조사", "회원가입", "거리(km)"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = thin

    for i, h in enumerate(results, start=1):
        row = [
            i, h["name"], h["sido"], h["sigungu"], h["type"],
            h["current_model"] or "-", h["current_maker"] or "-",
            "O" if h["is_member"] else "-", h["dist_km"] if h["dist_km"] is not None else "-",
        ]
        ws.append(row)
        for c in range(1, len(cols) + 1):
            ws.cell(row=i + 1, column=c).border = thin

    for col, width in zip("ABCDEFGHI", [6, 26, 10, 12, 12, 22, 16, 10, 10]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"songlim_map_{category}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.get("/hospital-search")
async def search_hospital(
    q: str = Query(..., min_length=2, description="검색할 병원명"),
    hospital_type: str = Query("일반", description="의원/병원/대학병원이면 심평원 실시간 검색, 동물병원이면 로컬 DB 검색"),
    db: AsyncSession = Depends(get_db),
):
    """회원가입 시 병원명 검색 자동완성 (DB구성요청서 3-1 참고).
    일반병원은 심평원 hospInfoServicev2를 실시간 조회, 동물병원은 로컬 hospitals 테이블(동물병원 마스터데이터 임포트 후)에서 검색.
    매칭 시 name/tel만 반환 — 주소 등 나머지 필드는 정책상 자동입력하지 않고 사용자가 직접 입력한다."""
    if hospital_type == "동물병원":
        rows = (
            await db.execute(
                select(Hospital)
                .where(Hospital.type == "동물병원", Hospital.name.ilike(f"%{q}%"))
                .limit(20)
            )
        ).scalars().all()
        return [
            {"name": h.name, "tel": None, "sido": h.sido, "sigungu": h.sigungu, "ykiho": h.ykiho}
            for h in rows
        ]

    if not HIRA_API_KEY:
        raise HTTPException(status_code=503, detail="심평원 API 키가 설정되지 않았습니다")
    try:
        resp = requests.get(
            HIRA_SEARCH_URL,
            params={"serviceKey": HIRA_API_KEY, "yadmNm": q, "numOfRows": "20", "pageNo": "1"},
            timeout=10,
        )
        root = ET.fromstring(resp.text)
    except Exception:
        raise HTTPException(status_code=502, detail="심평원 API 조회에 실패했습니다")

    result_code = root.findtext(".//resultCode")
    if result_code and result_code != "00":
        raise HTTPException(status_code=502, detail="심평원 API 오류: " + (root.findtext(".//resultMsg") or ""))

    results = []
    for item in root.findall(".//item"):
        results.append({
            "name": (item.findtext("yadmNm") or "").strip(),
            "tel": (item.findtext("telno") or "").strip() or None,
            "sido": (item.findtext("sidoCdNm") or "").strip(),
            "sigungu": (item.findtext("sgguCdNm") or "").strip(),
            "type": (item.findtext("clCdNm") or "").strip(),
            "ykiho": item.findtext("ykiho") or "",
        })
    return results


@router.get("/regions/sido")
async def get_sido_list(db: AsyncSession = Depends(get_db), user: User = Depends(require_staff)):
    q = select(Hospital.sido).where(Hospital.sido.is_not(None)).distinct().order_by(Hospital.sido)
    rows = (await db.execute(q)).scalars().all()
    return list(rows)


@router.get("/regions/sigungu")
async def get_sigungu_list(db: AsyncSession = Depends(get_db), sido: str = Query(...), user: User = Depends(require_staff)):
    q = (
        select(Hospital.sigungu)
        .where(Hospital.sido == sido, Hospital.sigungu.is_not(None))
        .distinct()
        .order_by(Hospital.sigungu)
    )
    rows = (await db.execute(q)).scalars().all()
    return list(rows)


@router.get("/hospital/{hospital_id}")
async def get_hospital_detail(hospital_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(require_staff)):
    """병원 상세 — 지도 마커/카드 클릭 시 우측 패널에 표시. 카테고리별 연도별 보유장비 이력 포함."""
    h = (await db.execute(select(Hospital).where(Hospital.id == hospital_id))).scalar_one_or_none()
    if not h:
        return {"error": "not found"}

    eq_rows = (
        await db.execute(
            select(Equipment)
            .where(Equipment.hospital_id == hospital_id)
            .order_by(Equipment.category, Equipment.year.desc())
        )
    ).scalars().all()

    by_category: dict[str, dict[int, list[dict]]] = {}
    for e in eq_rows:
        by_category.setdefault(e.category, {}).setdefault(e.year, []).append({
            "id": e.id, "manufacturer": e.manufacturer, "model": e.model, "eq_count": e.eq_count, "source": e.source,
        })

    yearly_by_category = {
        cat: [{"year": y, "models": models, "total": sum(m["eq_count"] for m in models)}
              for y, models in sorted(years.items(), reverse=True)]
        for cat, years in by_category.items()
    }

    return {
        "hospital": {
            "id": h.id, "name": h.name, "type": h.type,
            "sido": h.sido, "sigungu": h.sigungu, "address": h.address,
            "lat": h.lat, "lng": h.lng,
            "is_member": h.hospital_profile_id is not None,
        },
        "yearly_by_category": yearly_by_category,
    }


class ManualEquipmentIn(BaseModel):
    hospital_id: int
    category: str  # us | xray | ct | mri
    manufacturer: str
    model: str
    year: int
    eq_count: int = 1
    created_by: Optional[int] = None


@router.get("/equipment/catalog")
async def get_equipment_catalog(db: AsyncSession = Depends(get_db), category: str = "us", user: User = Depends(require_staff)):
    """수동 등록 시 제조사/장비 선택용 드롭다운 — 기존 임포트 데이터의 distinct 값 재사용."""
    q = (
        select(Equipment.manufacturer, Equipment.model)
        .where(Equipment.category == category, Equipment.model.is_not(None))
        .distinct()
    )
    rows = (await db.execute(q)).all()
    return [{"manufacturer": r[0], "model": r[1]} for r in rows]


@router.post("/equipment/manual")
async def register_manual_equipment(payload: ManualEquipmentIn, db: AsyncSession = Depends(get_db), user: User = Depends(require_staff)):
    """동물병원/2026 신규개원 등 공공데이터에 없는 병원의 장비를 직원이 직접 등록.
    등록 즉시 지도/병원정보에 매칭되도록 source='manual'로 저장."""
    eq = Equipment(
        hospital_id=payload.hospital_id,
        category=payload.category,
        year=payload.year,
        manufacturer=payload.manufacturer,
        model=payload.model,
        eq_count=payload.eq_count,
        source="manual",
        created_by=payload.created_by,
    )
    db.add(eq)
    await db.commit()
    await db.refresh(eq)
    return {"id": eq.id}


class ManualEquipmentUpdateIn(BaseModel):
    manufacturer: str
    model: str
    year: int
    eq_count: int = 1


@router.patch("/equipment/manual/{equipment_id}")
async def update_manual_equipment(equipment_id: int, payload: ManualEquipmentUpdateIn, db: AsyncSession = Depends(get_db), user: User = Depends(require_staff)):
    """수동 등록된 장비만 추후 수정 가능 (source='manual'인 행만 허용, 공공데이터 임포트 행은 보호)."""
    eq = (await db.execute(select(Equipment).where(Equipment.id == equipment_id))).scalar_one_or_none()
    if not eq:
        raise HTTPException(status_code=404, detail="장비를 찾을 수 없습니다")
    if eq.source != "manual":
        raise HTTPException(status_code=403, detail="공공데이터로 임포트된 장비는 수정할 수 없습니다")
    eq.manufacturer = payload.manufacturer
    eq.model = payload.model
    eq.year = payload.year
    eq.eq_count = payload.eq_count
    await db.commit()
    return {"id": eq.id}


@router.delete("/equipment/manual/{equipment_id}")
async def delete_manual_equipment(equipment_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(require_staff)):
    eq = (await db.execute(select(Equipment).where(Equipment.id == equipment_id))).scalar_one_or_none()
    if not eq:
        raise HTTPException(status_code=404, detail="장비를 찾을 수 없습니다")
    if eq.source != "manual":
        raise HTTPException(status_code=403, detail="공공데이터로 임포트된 장비는 삭제할 수 없습니다")
    await db.delete(eq)
    await db.commit()
    return {"ok": True}


# ────────────────────────────────────────────────────────
# 영업노트 — 병원별/날짜별/작성자별로 모아보기
# ────────────────────────────────────────────────────────
class SalesNoteIn(BaseModel):
    hospital_id: int
    visit_date: Optional[date] = None
    content: str


@router.post("/sales-notes")
async def create_sales_note(
    payload: SalesNoteIn, db: AsyncSession = Depends(get_db), user: User = Depends(require_staff)
):
    note = SalesNote(hospital_id=payload.hospital_id, user_id=user.id, visit_date=payload.visit_date, content=payload.content)
    db.add(note)
    await db.commit()
    await db.refresh(note)
    return {"id": note.id}


@router.get("/sales-notes")
async def list_sales_notes(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_staff),
    hospital_id: Optional[int] = None,
    user_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
):
    q = select(SalesNote, Hospital.name).join(Hospital, Hospital.id == SalesNote.hospital_id)
    if hospital_id:
        q = q.where(SalesNote.hospital_id == hospital_id)
    if user_id:
        q = q.where(SalesNote.user_id == user_id)
    if date_from:
        q = q.where(SalesNote.visit_date >= date_from)
    if date_to:
        q = q.where(SalesNote.visit_date <= date_to)
    q = q.order_by(SalesNote.visit_date.desc())
    rows = (await db.execute(q)).all()
    return [
        {
            "id": n.id, "hospital_id": n.hospital_id, "hospital_name": hosp_name, "user_id": n.user_id,
            "visit_date": n.visit_date, "content": n.content, "created_at": n.created_at,
        }
        for n, hosp_name in rows
    ]


# ────────────────────────────────────────────────────────
# 통계 — 제조사 매칭 기반, 영업지도와 동일 카테고리(us/xray/ct/mri) 사용
# ────────────────────────────────────────────────────────
@router.get("/stats/by-maker")
async def stats_by_maker(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_staff),
    category: str = "us",
    year: int = 2025,
    sido: Optional[str] = None,
):
    q = (
        select(Equipment.manufacturer, func.count(func.distinct(Equipment.hospital_id)))
        .join(Hospital, Hospital.id == Equipment.hospital_id)
        .where(Equipment.category == category, Equipment.year == year)
    )
    if sido:
        q = q.where(Hospital.sido == sido)
    q = q.group_by(Equipment.manufacturer)
    rows = (await db.execute(q)).all()
    return [{"maker": r[0] or "미상", "hospital_count": r[1]} for r in rows]
